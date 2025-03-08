import pandas as pd
import json
import requests
import os
import time
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def read_excel_data(file_path):
    """Read the Excel file with categories and topics related to China."""
    try:
        df = pd.read_excel(file_path)
        # Verify required columns exist
        if 'Category' not in df.columns or 'Topic' not in df.columns:
            print("Error: Excel file must contain 'Category' and 'Topic' columns.")
            return None
        return df
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return None

def generate_questions_and_answers(category, topic, num_questions=5):
    """Generate questions and answers for a category-topic pair using Ollama."""
    prompt = f"""
    Generate {num_questions} specific questions about {topic} related to {category} in China. 
    Each question should be detailed and focused on important aspects.
    For each question, also provide a comprehensive answer.
    Format the response as a JSON array with objects containing 'question' and 'answer' keys.
    """
    
    try:
        response = requests.post(
            'http://localhost:11434/api/generate',
            json={
                'model': 'llama3',  # You can change to another model available in your Ollama setup
                'prompt': prompt,
                'stream': False,
                'temperature': 0.7
            },
            timeout=120
        )
        
        if response.status_code == 200:
            result = response.json()
            response_text = result.get('response', '')
            
            # Extract JSON from the response
            try:
                # Find JSON array in the response
                start_idx = response_text.find('[')
                end_idx = response_text.rfind(']') + 1
                
                if start_idx != -1 and end_idx != -1:
                    json_str = response_text[start_idx:end_idx]
                    questions_answers = json.loads(json_str)
                    return questions_answers
                else:
                    # If JSON array markers not found, try parsing the whole response
                    questions_answers = json.loads(response_text)
                    return questions_answers
            except json.JSONDecodeError:
                # Fallback parsing approach
                lines = response_text.split('\n')
                questions_answers = []
                current_question = None
                current_answer = ""
                
                for line in lines:
                    if line.strip().startswith("Q") or line.strip().startswith("Question"):
                        # Save previous Q&A if exists
                        if current_question is not None:
                            questions_answers.append({
                                "question": current_question,
                                "answer": current_answer.strip()
                            })
                        # Start new question
                        parts = line.split(":", 1)
                        if len(parts) > 1:
                            current_question = parts[1].strip()
                            current_answer = ""
                    elif line.strip().startswith("A") or line.strip().startswith("Answer"):
                        parts = line.split(":", 1)
                        if len(parts) > 1:
                            current_answer = parts[1].strip()
                    elif current_question is not None:
                        current_answer += " " + line.strip()
                
                # Add the last Q&A
                if current_question is not None:
                    questions_answers.append({
                        "question": current_question,
                        "answer": current_answer.strip()
                    })
                
                return questions_answers[:num_questions]
        else:
            print(f"Error from Ollama API: {response.status_code} - {response.text}")
            return None
    except Exception as e:
        print(f"Error generating Q&A with Ollama: {e}")
        return None

def write_results_to_excel(input_file, output_file=None):
    """Process each category-topic pair and write results to a new Excel file."""
    if output_file is None:
        base, ext = os.path.splitext(input_file)
        output_file = f"{base}_with_qa{ext}"
    
    # Read the original data
    df = read_excel_data(input_file)
    if df is None:
        return
    
    # Create copies of the dataframe for questions and answers
    df_with_qa = df.copy()
    
    # Add separate columns for each question and answer
    for i in range(1, 6):
        df_with_qa[f'Question {i}'] = ''
        df_with_qa[f'Answer {i}'] = ''
    
    # Process each row
    for idx, row in df.iterrows():
        category = row['Category']
        topic = row['Topic']
        
        print(f"Processing: Category '{category}', Topic '{topic}'")
        
        qa_results = generate_questions_and_answers(category, topic)
        if qa_results:
            for i, qa in enumerate(qa_results[:5]):
                q_col = f'Question {i+1}'
                a_col = f'Answer {i+1}'
                # Store questions and answers in their respective columns
                df_with_qa.at[idx, q_col] = qa.get('question', '')
                df_with_qa.at[idx, a_col] = qa.get('answer', '')
        
        # Add a small delay to avoid overwhelming the Ollama API
        time.sleep(1)
    
    # Save the results
    df_with_qa.to_excel(output_file, index=False)
    
    # Adjust column widths for better readability
    wb = load_workbook(output_file)
    ws = wb.active
    
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        column_header = ws.cell(row=1, column=col).value
        
        # Set question columns wider
        if column_header and "Question" in column_header:
            ws.column_dimensions[col_letter].width = 50
        # Set answer columns even wider
        elif column_header and "Answer" in column_header:
            ws.column_dimensions[col_letter].width = 70
        else:
            ws.column_dimensions[col_letter].width = 20
    
    wb.save(output_file)
    
    print(f"Process complete. Results saved to {output_file}")
    print(f"The output file contains {len(df_with_qa)} rows with questions and answers in separate columns.")

if __name__ == "__main__":
    # Get the Excel file path from the user
    excel_file = input("Enter the path to your Excel file: ")
    
    # Check if Ollama is running
    try:
        response = requests.get('http://localhost:11434/api/version')
        if response.status_code == 200:
            print("Ollama is running.")
        else:
            print("Warning: Could not verify Ollama is running. Proceeding anyway.")
    except:
        print("Warning: Ollama service not detected. Please ensure Ollama is running.")
        proceed = input("Continue anyway? (y/n): ")
        if proceed.lower() != 'y':
            exit()
    
    # Process the file
    write_results_to_excel(excel_file)